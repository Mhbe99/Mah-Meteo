#!/usr/bin/env python3
"""
Script de rapport hebdomadaire des risques météo
Généré chaque lundi matin avec :
- Récapitulatif des alertes par jour/heure
- Graphiques et statistiques
- Envoi par email
"""

import json
import os
import datetime
from zoneinfo import ZoneInfo
from collections import defaultdict
import requests
from jose import jwt
from dotenv import load_dotenv
from meteo_saas.backend.email_alerts import _envoyer_email

# Charger les variables d'environnement
load_dotenv()
SENDER_EMAIL = os.getenv("SENDER_EMAIL")
GMAIL_PASSWORD = os.getenv("GMAIL_PASSWORD")
_receivers_raw = os.getenv("RECEIVER_EMAILS", "")
RECEIVER_EMAILS = [e.strip() for e in _receivers_raw.split(",") if e.strip()]
RENDER_URL = os.getenv("RENDER_URL", "https://mah-meteo.onrender.com")
JWT_SECRET = os.getenv("JWT_SECRET", "")
RENDER_API_TOKEN = os.getenv("RENDER_API_TOKEN", "")
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "") or SENDER_EMAIL
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "") or GMAIL_PASSWORD
SMTP_FROM = os.getenv("SMTP_FROM", "") or SENDER_EMAIL
ALLOW_LOCAL_REPORT_FALLBACK = os.getenv("ALLOW_LOCAL_REPORT_FALLBACK", "false").lower() == "true"
REPORT_REQUIRE_RENDER = os.getenv("REPORT_REQUIRE_RENDER", "true").lower() == "true"
REPORT_CLIENT_SCAN_MAX = int(os.getenv("REPORT_CLIENT_SCAN_MAX", "50"))
GITHUB_ACTIONS = os.getenv("GITHUB_ACTIONS", "false").lower() == "true"
PARIS_TZ = ZoneInfo("Europe/Paris")

# Paths
EXPORT_PATH = "exports"
RAPPORT_FILE = os.path.join(EXPORT_PATH, "rapport_hebdomadaire.xlsx")

# Emojis a exclure du rapport hebdomadaire (fichier Excel + libelles email)
_REPORT_EMOJI_TOKENS = [
    "🏣", "📍", "📊", "⚠️", "⚠", "📅", "🔥", "🗺️", "🗺", "📋",
    "❄️", "💨", "🌧️", "🌧", "☀️", "☀", "✅", "❌", "🌫️", "🌫",
    "🔴", "🟠", "⛔", "📎",
]


def _strip_report_emojis(value):
    """Retire les emojis des libelles pour garantir un rapport hebdo sans emoji."""
    if value is None:
        return ""
    text = str(value)
    for token in _REPORT_EMOJI_TOKENS:
        text = text.replace(token, "")
    text = text.replace("\ufe0f", "")
    return " ".join(text.split())


def _build_email_shell(title, subtitle, content_html, accent="#2c3e50"):
    now_str = datetime.datetime.now(PARIS_TZ).strftime("%d/%m/%Y à %H:%M")
    return f"""<!DOCTYPE html>
<html><head><meta charset=\"UTF-8\"><meta name=\"viewport\" content=\"width=device-width,initial-scale=1\"></head>
<body style=\"margin:0;padding:0;background:#eef1f5;font-family:'Segoe UI',Arial,sans-serif;\">
    <div style=\"max-width:680px;margin:24px auto;background:#fff;border:1px solid #e2e8f0;border-radius:8px;overflow:hidden;\">
        <div style=\"background:{accent};padding:16px 20px;\">
            <div style=\"font-size:11px;color:#cbd5e0;text-transform:uppercase;letter-spacing:1px;\">Mah Météo</div>
            <div style=\"font-size:18px;color:#fff;font-weight:700;margin-top:4px;\">{title}</div>
            <div style=\"font-size:12px;color:#e2e8f0;margin-top:2px;\">{subtitle}</div>
        </div>
        <div style=\"padding:18px 20px;color:#2d3748;font-size:13px;line-height:1.55;\">{content_html}</div>
        <div style=\"padding:12px 20px;background:#f7fafc;border-top:1px solid #e2e8f0;color:#718096;font-size:11px;\">
            Mah Météo GEODIS · Généré automatiquement · {now_str}
        </div>
    </div>
</body></html>"""

# --- Sites pour prévisions (mêmes sites que le dashboard) ---
SITES = {
    "Le Meux 🏣": {"lat": 49.378829, "lon": 2.750393},
    "Clairoix 🏣": {"lat": 49.4194, "lon": 2.8328},
}
VOISINS = {
    "Compiègne": {"lat": 49.4176, "lon": 2.8261},
    "Creil": {"lat": 49.2561, "lon": 2.4834},
    "Beauvais": {"lat": 49.4304, "lon": 2.0876},
    "Chantilly": {"lat": 49.1931, "lon": 2.4714},
    "Noyon": {"lat": 49.5786, "lon": 3.0017},
    "Senlis": {"lat": 49.2079, "lon": 2.5849},
    "Méru": {"lat": 49.2335, "lon": 2.1293},
    "Nogent-sur-Oise": {"lat": 49.2661, "lon": 2.4706},
    "Clermont": {"lat": 49.3763, "lon": 2.4151},
    "Montataire": {"lat": 49.2641, "lon": 2.4436},
    "Liancourt": {"lat": 49.3334, "lon": 2.4573},
    "Chaumont-en-Vexin": {"lat": 49.2761, "lon": 1.8759},
    "Formerie": {"lat": 49.65, "lon": 1.73},
    "Breteuil": {"lat": 49.633331, "lon": 2.3},
    "St Just en Chaussée": {"lat": 49.5, "lon": 2.43333},
    "Trosly-Breuil": {"lat": 49.400002, "lon": 2.96667},
    "Crépy-en-Valois": {"lat": 49.23333, "lon": 2.9},
    "Saint-Maximin": {"lat": 49.216671, "lon": 2.45},
    "Nanteuil-le-Haudouin": {"lat": 49.133331, "lon": 2.8},
}

def _parse_iso(ts: str):
    """Parse un timestamp ISO en tolérant le suffixe Z."""
    if not ts:
        return None
    try:
        return datetime.datetime.fromisoformat(ts.replace("Z", "+00:00"))
    except Exception:
        return None


def _charger_historique_local():
    """Fallback local quand l'API Render n'est pas accessible."""
    archive_file = os.path.join(EXPORT_PATH, "alertes_historique.json")
    if not os.path.exists(archive_file):
        return []
    try:
        with open(archive_file, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        print(f"ℹ️ Fallback local: {len(data)} alertes chargées depuis {archive_file}")
        return data
    except Exception as e:
        print(f"⚠️ Impossible de lire le fallback local: {e}")
        return []


def _build_local_client_token(client_id):
    """Construit un JWT local pour interroger /api/alertes/{client_id} sans /api/service/*."""
    if not JWT_SECRET:
        return ""
    payload = {
        "client_id": client_id,
        "username": f"report-client-{client_id}",
        "exp": datetime.datetime.utcnow() + datetime.timedelta(hours=2),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")


def _fetch_alertes_for_client_id(client_id, token):
    """Récupère les alertes d'un client via token Bearer applicatif."""
    ra = requests.get(
        f"{RENDER_URL}/api/alertes/{client_id}",
        params={"limit": 500},
        headers={"Authorization": f"Bearer {token}"},
        timeout=15,
    )
    return ra


def _collect_alertes_direct_jwt_scan(max_client_id):
    """Fallback Render: scan des client_id actifs en JWT local quand /api/service/* renvoie 403."""
    alertes = []
    active_client_ids = []

    for cid in range(1, max_client_id + 1):
        token_c = _build_local_client_token(cid)
        if not token_c:
            break
        try:
            ra = _fetch_alertes_for_client_id(cid, token_c)
        except Exception as e:
            print(f"⚠️ Alertes client {cid} indisponibles: {e}")
            continue

        if ra.status_code == 200:
            active_client_ids.append(cid)
            for a in ra.json():
                ts = a.get("timestamp") or a.get("created_at") or ""
                ts_dt = _parse_iso(ts)
                alertes.append({
                    "timestamp": ts,
                    "date": ts_dt.strftime("%Y-%m-%d") if ts_dt else (ts[:10] if ts else ""),
                    "heure": ts_dt.strftime("%H:%M") if ts_dt else (ts[11:16] if len(ts) > 15 else ""),
                    "jour_semaine": ts_dt.strftime("%A") if ts_dt else "",
                    "zone": a.get("zone_name", ""),
                    "risques": a.get("message", ""),
                    "temp": "",
                    "wind": "",
                    "rain": "",
                })
            continue

        if ra.status_code in (401, 403, 404):
            continue

        print(f"⚠️ Alertes client {cid} indisponibles: HTTP {ra.status_code}")

    return alertes, active_client_ids


def charger_historique():
    """Charge l'historique Render; fallback local uniquement si explicitement autorise."""
    service_headers = {}
    service_auth_secret = (RENDER_API_TOKEN or JWT_SECRET or "").strip()
    if service_auth_secret:
        # Compat: certaines versions lisent X-Service-Secret/X-Service-Key, d'autres Authorization Bearer.
        service_headers = {
            "X-Service-Secret": service_auth_secret,
            "X-Service-Key": service_auth_secret,
            "Authorization": f"Bearer {service_auth_secret}",
        }

    try:
        rc = requests.get(
            f"{RENDER_URL}/api/service/clients",
            headers=service_headers,
            timeout=15,
        )
        alertes = []
        if rc.status_code == 200:
            clients = rc.json().get("clients", [])

            for client in clients:
                cid = client.get("id") or client.get("client_id")
                if not cid:
                    continue

                # Token par client via endpoint service sécurisé
                rt = requests.get(
                    f"{RENDER_URL}/api/service/token",
                    params={"client_id": cid},
                    headers=service_headers,
                    timeout=10,
                )
                if rt.status_code != 200:
                    print(f"⚠️ Token client {cid} indisponible: HTTP {rt.status_code}")
                    continue

                td = rt.json()
                token_c = td.get("access_token") or td.get("token", "")
                if not token_c:
                    print(f"⚠️ Token vide pour client {cid}")
                    continue

                ra = _fetch_alertes_for_client_id(cid, token_c)
                if ra.status_code != 200:
                    print(f"⚠️ Alertes client {cid} indisponibles: HTTP {ra.status_code}")
                    continue

                for a in ra.json():
                    ts = a.get("timestamp") or a.get("created_at") or ""
                    ts_dt = _parse_iso(ts)
                    alertes.append({
                        "timestamp": ts,
                        "date": ts_dt.strftime("%Y-%m-%d") if ts_dt else (ts[:10] if ts else ""),
                        "heure": ts_dt.strftime("%H:%M") if ts_dt else (ts[11:16] if len(ts) > 15 else ""),
                        "jour_semaine": ts_dt.strftime("%A") if ts_dt else "",
                        "zone": a.get("zone_name", ""),
                        "risques": a.get("message", ""),
                        "temp": "",
                        "wind": "",
                        "rain": "",
                    })
        elif rc.status_code == 403:
            if GITHUB_ACTIONS:
                print("[RAPPORT] /api/service/* refuse le secret (403) en CI: verifier les secrets runtime Render/GitHub")
            else:
                print(
                    "[RAPPORT] /api/service/* refuse le secret (403) en local: "
                    "mismatch local possible, non bloquant si le workflow GitHub passe"
                )
            print("[RAPPORT] Bascule en scan JWT direct")
            alertes, active_ids = _collect_alertes_direct_jwt_scan(REPORT_CLIENT_SCAN_MAX)
            print(f"[RAPPORT] Clients actifs détectés via JWT direct: {len(active_ids)}")
        else:
            rc.raise_for_status()

        if alertes:
            print(f"[RAPPORT] Source : API Render ({len(alertes)} alertes)")
            return alertes, "API Render (production)"

        print("[RAPPORT] API Render sans alerte exploitable")
        if ALLOW_LOCAL_REPORT_FALLBACK:
            local_alertes = _charger_historique_local()
            print(f"[RAPPORT] Source : fallback local ({len(local_alertes)} alertes)")
            return local_alertes, "Fallback local"

        print("[RAPPORT] Fallback local désactivé (ALLOW_LOCAL_REPORT_FALLBACK=false)")
        return [], "API Render vide (production)"
    except Exception as e:
        print(f"[RAPPORT] Erreur API Render : {e}")
        if ALLOW_LOCAL_REPORT_FALLBACK:
            local_alertes = _charger_historique_local()
            print(f"[RAPPORT] Source : fallback local ({len(local_alertes)} alertes)")
            return local_alertes, "Fallback local"

        print("[RAPPORT] Fallback local désactivé (ALLOW_LOCAL_REPORT_FALLBACK=false)")
        return [], "API Render indisponible"

def get_semaine_precedente():
    """Retourner la date de début et fin de la semaine précédente (lundi-dimanche)"""
    today = datetime.datetime.now(PARIS_TZ)
    # Lundi de la semaine précédente
    jours_avant = (today.weekday() + 7) % 7
    lundi = today - datetime.timedelta(days=jours_avant + 7)
    dimanche = lundi + datetime.timedelta(days=6)
    return lundi.date(), dimanche.date()

def filtrer_semaine(alertes):
    """Filtrer les alertes de la semaine précédente"""
    lundi, dimanche = get_semaine_precedente()
    alertes_semaine = []
    for alerte in alertes:
        try:
            date_alerte = datetime.datetime.fromisoformat(alerte["timestamp"].replace("Z", "+00:00")).date()
            if lundi <= date_alerte <= dimanche:
                alertes_semaine.append(alerte)
        except:
            pass
    return alertes_semaine

def generer_statistiques(alertes_semaine):
    """Générer les statistiques des alertes"""
    stats = {
        "total_alertes": len(alertes_semaine),
        "zones": defaultdict(int),
        "risques": defaultdict(int),
        "par_jour": defaultdict(int),
        "par_heure": defaultdict(int),
        "par_zone_heure": defaultdict(lambda: defaultdict(int))
    }
    
    for alerte in alertes_semaine:
        zone = alerte.get("zone", "Inconnue")
        jour = alerte.get("jour_semaine", "?")
        heure = alerte.get("heure", "?")
        risques_str = alerte.get("risques", "")
        
        stats["zones"][zone] += 1
        stats["par_jour"][jour] += 1
        stats["par_heure"][heure] += 1
        stats["par_zone_heure"][zone][heure] += 1
        
        # Comptabiliser les risques individuels
        for risque in ["Verglas", "Vent fort", "Alerte pluie", "UV fort"]:
            if risque in risques_str:
                stats["risques"][risque] += 1
    
    return stats


def get_risk_icons(temp, wind, rain, uv):
    """Petit heuristique pour identifier les risques (copie légère)."""
    risk = []
    try:
        t = float(temp)
        r = float(rain)
        if t < 1 and r > 0 and datetime.datetime.now().month in [11, 12, 1, 2]:
            risk.append("❄️ Verglas")
    except:
        pass
    if wind is not None and float(wind) > 40:
        risk.append("💨 Vent fort")
    if rain is not None and float(rain) > 5:
        risk.append("🌧️ Alerte pluie")
    if uv is not None and float(uv) >= 8:
        risk.append("🔥 UV fort")
    return " | ".join(risk) if risk else "✅ RAS"


def fetch_previsions_5j(sites, jours=5):
    """Récupère les prévisions journalières (jours prochains) pour les sites fournis.

    Retourne dict: {site: [ {jour,tmin,tmax,pluie,uv,risk}, ... ] }
    """
    result = {site: [] for site in sites}
    names = list(sites.keys())
    coords = list(sites.values())
    lats = ",".join(str(c.get("lat")) for c in coords)
    lons = ",".join(str(c.get("lon")) for c in coords)
    batch_url = (
        f"https://api.open-meteo.com/v1/forecast?latitude={lats}&longitude={lons}"
        f"&daily=temperature_2m_max,temperature_2m_min,precipitation_sum,uv_index_max&timezone=auto"
    )

    batch_data = None
    try:
        r = requests.get(batch_url, timeout=25)
        r.raise_for_status()
        raw = r.json()
        batch_data = raw if isinstance(raw, list) else [raw]
    except Exception as e:
        print(f"⚠️ Erreur batch prévisions 5j: {e} — fallback unitaire")

    for i, site in enumerate(names):
        data = None
        if batch_data and i < len(batch_data):
            data = batch_data[i]
        else:
            coord = sites[site]
            lat, lon = coord.get("lat"), coord.get("lon")
            url = (
                f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}"
                f"&daily=temperature_2m_max,temperature_2m_min,precipitation_sum,uv_index_max&timezone=auto"
            )
            try:
                rr = requests.get(url, timeout=20)
                rr.raise_for_status()
                data = rr.json()
            except Exception as e:
                print(f"⚠️ Erreur récupération prévisions pour {site}: {e}")
                continue

        try:
            days = data.get("daily", {})
            times = days.get("time", [])
            for j in range(min(len(times), jours)):
                date = datetime.datetime.strptime(times[j], "%Y-%m-%d").strftime("%a %d/%m")
                tmin = days.get('temperature_2m_min', [None])[j]
                tmax = days.get('temperature_2m_max', [None])[j]
                pluie = days.get('precipitation_sum', [0])[j]
                uv = days.get('uv_index_max', [0])[j]
                risk = get_risk_icons(tmin if tmin is not None else 0, 0, pluie, uv)
                result[site].append({
                    "jour": date,
                    "tmin": f"{tmin}\u00b0C" if tmin is not None else "N/A",
                    "tmax": f"{tmax}\u00b0C" if tmax is not None else "N/A",
                    "pluie": f"{pluie} mm",
                    "uv": uv,
                    "risk": risk,
                })
        except Exception as e:
            print(f"⚠️ Parsing prévisions impossible pour {site}: {e}")
    return result

def _build_zone_card(zone_name, jours, is_site=False):
    """Génère une carte HTML prévision pour une zone (style dashboard)."""
    icon = "🏣" if is_site else "📍"
    bg = "#edf7f1" if is_site else "#f7fafc"

    days_html = ""
    for j in jours:
        # Couleur du badge risque
        risk_text = j['risk']
        if "❄️" in risk_text or "Verglas" in risk_text:
            rbg, rcol = "#eef4fb", "#1a4a7a"
        elif "💨" in risk_text or "Vent" in risk_text:
            rbg, rcol = "#fdf6ec", "#b8660a"
        elif "🌧️" in risk_text or "pluie" in risk_text.lower():
            rbg, rcol = "#ebf8ff", "#2b6cb0"
        elif "🔥" in risk_text or "UV" in risk_text:
            rbg, rcol = "#fdf0ee", "#c0392b"
        else:
            rbg, rcol = "#edf7f1", "#1a6b3a"

        days_html += f"""
        <td style="padding:10px 6px;text-align:center;border-right:1px solid #e2e8f0;vertical-align:top;width:20%;">
          <div style="font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;color:#a0aec0;margin-bottom:6px;">{j['jour']}</div>
          <div style="font-size:13px;margin-bottom:4px;"><span style="color:#c0392b;font-weight:500;">{j['tmax']}</span> <span style="color:#ccc;">/</span> <span style="color:#2b6cb0;">{j['tmin']}</span></div>
          <div style="font-size:11px;color:#a0aec0;margin-bottom:3px;">🌧 {j['pluie']}</div>
          <div style="font-size:11px;color:#a0aec0;margin-bottom:6px;">☀️ UV {j['uv']}</div>
          <div style="display:inline-block;padding:2px 8px;border-radius:3px;font-size:10px;font-weight:500;background:{rbg};color:{rcol};">{risk_text}</div>
        </td>"""

    return f"""
    <div style="background:#fff;border:1px solid #dce1e8;border-radius:6px;margin-bottom:10px;overflow:hidden;">
      <div style="padding:8px 14px;background:{bg};border-bottom:1px solid #dce1e8;font-weight:600;font-size:13px;color:#2d3748;">
        {icon} {zone_name}
      </div>
      <table style="width:100%;border-collapse:collapse;">
        <tr>{days_html}</tr>
      </table>
    </div>"""


def fetch_pollution_actuelle():
    """Récupère les données AQI actuelles des sites pour affichage dans le rapport."""
    try:
        token_r = requests.get(
            f"{RENDER_URL}/api/service/token?client_id=1",
            headers={"X-Service-Key": JWT_SECRET},
            timeout=5
        )
        if token_r.status_code != 200:
            return None
        
        token = token_r.json().get("access_token") or token_r.json().get("token", "")
        if not token:
            return None
        
        meteo_r = requests.get(
            f"{RENDER_URL}/api/meteo/1",
            headers={"Authorization": f"Bearer {token}"},
            timeout=5
        )
        if meteo_r.status_code != 200:
            return None
        
        zones = meteo_r.json()
        sites_pollution = [z for z in zones if z.get("type") == "site" and z.get("aqi") is not None and z.get("aqi") >= 40]
        
        return sites_pollution if sites_pollution else None
    except Exception as e:
        print(f"[RAPPORT] Erreur récupération pollution: {e}")
        return None


def build_pollution_section():
    """Crée la section HTML pour la pollution si des données sont disponibles."""
    pollution_data = fetch_pollution_actuelle()
    if not pollution_data:
        return ""
    
    rows_html = ""
    for z in pollution_data:
        aqi = z.get("aqi", 0)
        label = z.get("pollution_label", "Inconnu")
        
        # Couleur selon seuil
        if aqi >= 80:
            color = "#7b341e"
            emoji = "⛔"
        elif aqi >= 60:
            color = "#e53e3e"
            emoji = "🔴"
        else:
            color = "#dd6b20"
            emoji = "🟠"
        
        rows_html += f"""<tr style="border-bottom:1px solid #e2e8f0;">
          <td style="padding:8px 12px;font-weight:600;color:#1a202c;">{z.get('name', 'Inconnu')}</td>
          <td style="padding:8px 12px;text-align:center;font-weight:700;color:{color};font-size:18px;">{round(aqi)}</td>
          <td style="padding:8px 12px;text-align:center;color:{color};">{emoji} {label}</td>
        </tr>"""
    
        return f"""<div style="background:#fffbeb;border-left:4px solid #dd6b20;padding:14px;border-radius:6px;margin:14px 0;">
            <div style="font-weight:700;color:#dd6b20;margin-bottom:10px;font-size:14px;">Qualite de l'air actuellement</div>
      <table style="width:100%;border-collapse:collapse;font-size:12px;">
        <thead>
          <tr style="background:#fef5e7;">
            <th style="padding:8px 12px;text-align:left;border-bottom:1px solid #dd6b20;color:#2d3748;font-weight:600;">Site</th>
            <th style="padding:8px 12px;text-align:center;border-bottom:1px solid #dd6b20;color:#2d3748;font-weight:600;">Pollution (AQI)</th>
            <th style="padding:8px 12px;text-align:center;border-bottom:1px solid #dd6b20;color:#2d3748;font-weight:600;">Niveau</th>
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>
      <p style="font-size:11px;color:#92400e;margin:8px 0 0 0;">Échelle: 40-59 Modéré · 60-79 Mauvais · 80+ Très mauvais</p>
    </div>"""


def envoyer_rapport_email(semaine_debut, semaine_fin, stats, rapport_file,
                          previsions_sites=None, previsions_voisins=None,
                          dry_run=False, source_donnees="API Render (production)"):
    """Envoyer le rapport par email"""
    
    # --- KPI cards ---
    kpi_html = f"""
    <div style="display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap;">
        <div style="flex:1;min-width:120px;background:#ebf8ff;border-left:4px solid #3182ce;border-radius:4px;padding:12px 16px;">
            <div style="font-size:22px;font-weight:700;color:#2b6cb0;">{stats['total_alertes']}</div>
            <div style="font-size:12px;color:#4a5568;margin-top:2px;">Alertes totales</div>
        </div>
        <div style="flex:1;min-width:120px;background:#fff5f5;border-left:4px solid #e53e3e;border-radius:4px;padding:12px 16px;">
            <div style="font-size:22px;font-weight:700;color:#c53030;">{len(stats['zones'])}</div>
            <div style="font-size:12px;color:#4a5568;margin-top:2px;">Zones affectées</div>
        </div>
        <div style="flex:1;min-width:120px;background:#fffaf0;border-left:4px solid #d69e2e;border-radius:4px;padding:12px 16px;">
            <div style="font-size:22px;font-weight:700;color:#b7791f;">{len(stats['risques'])}</div>
            <div style="font-size:12px;color:#4a5568;margin-top:2px;">Types de risques</div>
        </div>
    </div>
    """

    # --- Tableau risques ---
    risques_rows = ""
    for risque, count in sorted(stats['risques'].items(), key=lambda x: x[1], reverse=True):
        risques_rows += f"""
        <tr>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;font-size:13px;">{risque}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;font-size:13px;font-weight:600;color:#e53e3e;">{count}</td>
        </tr>"""

    # --- Tableau zones ---
    zones_rows = ""
    for zone, count in sorted(stats['zones'].items(), key=lambda x: x[1], reverse=True):
        zones_rows += f"""
        <tr>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;font-size:13px;">{zone}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;font-size:13px;font-weight:600;color:#3182ce;">{count}</td>
        </tr>"""

    # --- Prévisions visuelles (cartes par zone comme le dashboard) ---
    prev_sites_html = ""
    if previsions_sites:
        for site, jours in previsions_sites.items():
            if jours:
                prev_sites_html += _build_zone_card(site, jours, is_site=True)

    prev_voisins_html = ""
    if previsions_voisins:
        for voisin, jours in previsions_voisins.items():
            if jours:
                prev_voisins_html += _build_zone_card(voisin, jours, is_site=False)

    # Commentaire de source visible pour éviter les ambiguïtés de données en production.
    content_html = f"""
    {kpi_html}
    {build_pollution_section()}

    <h3 style="margin:20px 0 10px 0;font-size:14px;color:#2d3748;border-bottom:2px solid #f0f4f8;padding-bottom:6px;">Risques par type</h3>
    <table style="width:100%;border-collapse:collapse;border:1px solid #e2e8f0;border-radius:4px;overflow:hidden;">
        <thead>
            <tr style="background:#f7fafc;">
                <th style="padding:8px 12px;text-align:left;font-size:12px;color:#4a5568;">Type de risque</th>
                <th style="padding:8px 12px;text-align:left;font-size:12px;color:#4a5568;">Occurrences</th>
            </tr>
        </thead>
        <tbody>{risques_rows if risques_rows else '<tr><td colspan="2" style="padding:10px 12px;font-size:13px;color:#a0aec0;">Aucun risque détecté</td></tr>'}</tbody>
    </table>

    <h3 style="margin:20px 0 10px 0;font-size:14px;color:#2d3748;border-bottom:2px solid #f0f4f8;padding-bottom:6px;">Zones les plus affectees</h3>
    <table style="width:100%;border-collapse:collapse;border:1px solid #e2e8f0;border-radius:4px;overflow:hidden;">
        <thead>
            <tr style="background:#f7fafc;">
                <th style="padding:8px 12px;text-align:left;font-size:12px;color:#4a5568;">Zone</th>
                <th style="padding:8px 12px;text-align:left;font-size:12px;color:#4a5568;">Alertes</th>
            </tr>
        </thead>
        <tbody>{zones_rows if zones_rows else '<tr><td colspan="2" style="padding:10px 12px;font-size:13px;color:#a0aec0;">Aucune zone affectée</td></tr>'}</tbody>
    </table>

    <h3 style="margin:20px 0 10px 0;font-size:14px;color:#2d3748;border-bottom:2px solid #f0f4f8;padding-bottom:6px;">Previsions 5 jours - Sites</h3>
    {prev_sites_html if prev_sites_html else '<p style="font-size:13px;color:#a0aec0;">Aucune prévision disponible</p>'}

    <h3 style="margin:20px 0 10px 0;font-size:14px;color:#2d3748;border-bottom:2px solid #f0f4f8;padding-bottom:6px;">Previsions 5 jours - Villes voisines</h3>
    {prev_voisins_html if prev_voisins_html else '<p style="font-size:13px;color:#a0aec0;">Aucune prévision disponible</p>'}

    <p style="margin-top:20px;font-size:12px;color:#a0aec0;">Rapport Excel detaille en piece jointe.</p>
    <p style="color:#888;font-size:11px">Source : {source_donnees}</p>
    """
    html_body = _build_email_shell(
        title="Rapport hebdomadaire des risques meteo",
        subtitle=f"Semaine du {semaine_debut} au {semaine_fin}",
        content_html=content_html,
        accent="#2c3e50",
    )

    subject = f"Rapport Hebdomadaire Meteo - {semaine_debut} au {semaine_fin}"
    attachments = []
    if os.path.exists(rapport_file):
        try:
            with open(rapport_file, "rb") as attachment:
                attachments.append({
                    "name": os.path.basename(rapport_file),
                    "content": attachment.read(),
                })
        except Exception as e:
            print(f"⚠️ Erreur lecture pièce jointe: {e}")
    
    # Si dry_run, ne pas envoyer : sauvegarder HTML et le message MIME dans exports/
    if dry_run:
        try:
            os.makedirs(EXPORT_PATH, exist_ok=True)
            html_path = os.path.join(EXPORT_PATH, "test_email.html")
            eml_path = os.path.join(EXPORT_PATH, "test_email.eml")
            with open(html_path, "w", encoding="utf-8") as fh:
                fh.write(html_body)
            with open(eml_path, "w", encoding="utf-8") as fh:
                fh.write(f"Subject: {subject}\nTo: {', '.join(RECEIVER_EMAILS)}\n\nDry run uniquement. Source: {source_donnees}\n")
            print(f"✅ Dry-run : email sauvegardé dans {html_path} et {eml_path}")
            return True
        except Exception as e:
            print(f"❌ Erreur sauvegarde email dry-run: {e}")
            return False

    # Le rapport passe désormais par le transport central unique.
    return _envoyer_email(subject, html_body, RECEIVER_EMAILS, attachments=attachments)


def generer_excel(alertes_semaine, stats):
    """Générer un fichier Excel vraiment opérationnel avec 4 onglets et graphiques avancés"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # ===== ONGLET 1: SYNTHÈSE OPÉRATIONNELLE =====
        ws_synthese = wb.create_sheet("Synthese", 0)
        
        # En-tête
        ws_synthese.merge_cells('A1:F1')
        header = ws_synthese['A1']
        header.value = "RAPPORT OPERATIONNEL - METEO & ALERTES"
        header.font = Font(size=16, bold=True, color="FFFFFF")
        header.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header.alignment = Alignment(horizontal="center", vertical="center")
        ws_synthese.row_dimensions[1].height = 30
        
        ws_synthese['A2'] = "Rapport généré:"
        ws_synthese['B2'] = datetime.datetime.now(PARIS_TZ).strftime("%d/%m/%Y à %H:%M")
        ws_synthese['A2'].font = Font(bold=True)
        
        # KPIs
        ws_synthese['A4'] = "KPIs SEMAINE"
        ws_synthese['A4'].font = Font(size=12, bold=True, color="FFFFFF")
        ws_synthese['A4'].fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        ws_synthese.merge_cells('A4:F4')
        
        kpi_data = [
            ["Total Alertes", stats['total_alertes'], "#3498db"],
            ["Zones Affectées", len(stats['zones']), "#e74c3c"],
            ["Types Risques", len(stats['risques']), "#f39c12"],
            ["Jours Actifs", len(stats['par_jour']), "#27ae60"],
        ]
        
        row = 5
        for kpi_label, kpi_value, kpi_color in kpi_data:
            cell_label = ws_synthese[f'A{row}']
            cell_label.value = kpi_label
            cell_label.font = Font(bold=True, size=11)
            cell_label.fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
            
            cell_value = ws_synthese[f'B{row}']
            cell_value.value = kpi_value
            cell_value.font = Font(size=14, bold=True, color="FFFFFF")
            cell_value.fill = PatternFill(start_color=kpi_color[1:], end_color=kpi_color[1:], fill_type="solid")
            cell_value.alignment = Alignment(horizontal="center")
            
            row += 1
        
        # Risques
        ws_synthese['A11'] = "RISQUES DETECTES"
        ws_synthese['A11'].font = Font(size=12, bold=True, color="FFFFFF")
        ws_synthese['A11'].fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
        ws_synthese.merge_cells('A11:B11')
        
        ws_synthese['A12'] = "Type"
        ws_synthese['B12'] = "Occurrences"
        for col in ['A', 'B']:
            ws_synthese[f'{col}12'].font = Font(bold=True, color="FFFFFF")
            ws_synthese[f'{col}12'].fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        
        risk_row = 13
        risk_last_row = risk_row
        for risque, count in sorted(stats['risques'].items(), key=lambda x: x[1], reverse=True):
            ws_synthese[f'A{risk_row}'] = _strip_report_emojis(risque)
            ws_synthese[f'B{risk_row}'] = count
            ws_synthese[f'B{risk_row}'].alignment = Alignment(horizontal="center")
            if risk_row % 2 == 0:
                for col in ['A', 'B']:
                    ws_synthese[f'{col}{risk_row}'].fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
            risk_row += 1
            risk_last_row = risk_row - 1
        
        # Graphique risques
        if risk_last_row >= 13:
            chart_risk = BarChart()
            chart_risk.type = "col"
            chart_risk.title = "Distribution des risques (semaine)"
            data_risk = Reference(ws_synthese, min_col=2, min_row=12, max_row=risk_last_row)
            cats_risk = Reference(ws_synthese, min_col=1, min_row=13, max_row=risk_last_row)
            chart_risk.add_data(data_risk, titles_from_data=True)
            chart_risk.set_categories(cats_risk)
            chart_risk.height = 8
            chart_risk.width = 15
            chart_risk.dLbls = DataLabelList()
            chart_risk.dLbls.showCatName = True
            chart_risk.dLbls.showVal = True
            ws_synthese.add_chart(chart_risk, "D11")
        
        # Alertes par jour
        ws_synthese['A25'] = "ALERTES PAR JOUR"
        ws_synthese['A25'].font = Font(size=12, bold=True, color="FFFFFF")
        ws_synthese['A25'].fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
        ws_synthese.merge_cells('A25:B25')
        
        ws_synthese['A26'] = "Jour"
        ws_synthese['B26'] = "Alertes"
        for col in ['A', 'B']:
            ws_synthese[f'{col}26'].font = Font(bold=True, color="FFFFFF")
            ws_synthese[f'{col}26'].fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        
        ordered_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday",
                       "lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi", "dimanche"]
        used_days = set()
        day_row = 27
        day_last_row = day_row
        for d in ordered_days:
            count = stats["par_jour"].get(d, 0)
            if count > 0:
                ws_synthese[f'A{day_row}'] = d
                ws_synthese[f'B{day_row}'] = count
                ws_synthese[f'B{day_row}'].alignment = Alignment(horizontal="center")
                if day_row % 2 == 0:
                    for col in ['A', 'B']:
                        ws_synthese[f'{col}{day_row}'].fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
                day_row += 1
                day_last_row = day_row - 1
                used_days.add(d)
        
        for d, count in sorted(stats["par_jour"].items()):
            if d not in used_days and count > 0:
                ws_synthese[f'A{day_row}'] = d
                ws_synthese[f'B{day_row}'] = count
                ws_synthese[f'B{day_row}'].alignment = Alignment(horizontal="center")
                if day_row % 2 == 0:
                    for col in ['A', 'B']:
                        ws_synthese[f'{col}{day_row}'].fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
                day_row += 1
                day_last_row = day_row - 1
        
        # Graphique tendance
        if day_last_row >= 27:
            chart_days = LineChart()
            chart_days.title = "Tendance des alertes (semaine)"
            chart_days.y_axis.title = "Nombre d'alertes"
            data_days = Reference(ws_synthese, min_col=2, min_row=26, max_row=day_last_row)
            cats_days = Reference(ws_synthese, min_col=1, min_row=27, max_row=day_last_row)
            chart_days.add_data(data_days, titles_from_data=True)
            chart_days.set_categories(cats_days)
            chart_days.height = 8
            chart_days.width = 15
            ws_synthese.add_chart(chart_days, "D25")
        
        ws_synthese.column_dimensions['A'].width = 20
        ws_synthese.column_dimensions['B'].width = 15
        
        # ===== ONGLET 2: HEATMAP ZONE-HEURE =====
        ws_heatmap = wb.create_sheet("Heatmap", 1)
        
        ws_heatmap['A1'] = "Zone"
        for h in range(24):
            heure_col = get_column_letter(h + 2)
            ws_heatmap[f'{heure_col}1'] = f"{h:02d}:00"
        ws_heatmap[f'{get_column_letter(26)}1'] = "TOTAL"
        
        for col_idx in range(1, 27):
            col_letter = get_column_letter(col_idx)
            ws_heatmap[f'{col_letter}1'].font = Font(bold=True, color="FFFFFF")
            ws_heatmap[f'{col_letter}1'].fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
            ws_heatmap[f'{col_letter}1'].alignment = Alignment(horizontal="center")
        
        zone_row = 2
        zone_intensities = [max(v.values()) if v else 0 for v in stats['par_zone_heure'].values()]
        max_intensity = max(zone_intensities, default=1)
        for zone in sorted(stats['par_zone_heure'].keys()):
            ws_heatmap[f'A{zone_row}'] = zone
            ws_heatmap[f'A{zone_row}'].font = Font(bold=True)
            
            total_zone = 0
            for h in range(24):
                heure = f"{h:02d}:00"
                count = stats['par_zone_heure'][zone].get(heure, 0)
                col_letter = get_column_letter(h + 2)
                ws_heatmap[f'{col_letter}{zone_row}'] = count
                ws_heatmap[f'{col_letter}{zone_row}'].alignment = Alignment(horizontal="center")
                
                if count > 0:
                    intensity = count / max_intensity
                    if intensity > 0.7:
                        color = "c0392b"
                    elif intensity > 0.4:
                        color = "e74c3c"
                    elif intensity > 0.1:
                        color = "f5b7b1"
                    else:
                        color = "fadbd8"
                    ws_heatmap[f'{col_letter}{zone_row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    if intensity > 0.5:
                        ws_heatmap[f'{col_letter}{zone_row}'].font = Font(color="FFFFFF")
                
                total_zone += count
            
            col_total = get_column_letter(26)
            ws_heatmap[f'{col_total}{zone_row}'] = total_zone
            ws_heatmap[f'{col_total}{zone_row}'].font = Font(bold=True, color="FFFFFF")
            ws_heatmap[f'{col_total}{zone_row}'].fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
            
            zone_row += 1
        
        ws_heatmap.column_dimensions['A'].width = 20
        for h in range(24):
            col_letter = get_column_letter(h + 2)
            ws_heatmap.column_dimensions[col_letter].width = 8
        ws_heatmap.column_dimensions[get_column_letter(26)].width = 10
        
        # ===== ONGLET 3: TOP ZONES =====
        ws_zones = wb.create_sheet("Top Zones", 2)
        
        ws_zones['A1'] = "Zone"
        ws_zones['B1'] = "Alertes"
        ws_zones['C1'] = "% du total"
        
        for col in ['A', 'B', 'C']:
            ws_zones[f'{col}1'].font = Font(bold=True, color="FFFFFF")
            ws_zones[f'{col}1'].fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
            ws_zones[f'{col}1'].alignment = Alignment(horizontal="center")
        
        total_all = sum(stats['zones'].values())
        zone_row = 2
        zone_last_row = zone_row
        
        for zone, count in sorted(stats["zones"].items(), key=lambda x: x[1], reverse=True):
            ws_zones[f'A{zone_row}'] = _strip_report_emojis(zone)
            ws_zones[f'B{zone_row}'] = count
            ws_zones[f'C{zone_row}'] = round(count / total_all * 100, 1) if total_all > 0 else 0
            
            ws_zones[f'B{zone_row}'].alignment = Alignment(horizontal="center")
            ws_zones[f'C{zone_row}'].alignment = Alignment(horizontal="center")
            
            if zone_row % 2 == 0:
                for col in ['A', 'B', 'C']:
                    ws_zones[f'{col}{zone_row}'].fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
            
            zone_row += 1
            zone_last_row = zone_row - 1
        
        # Graphique top zones
        if zone_last_row >= 2:
            chart_zones = BarChart()
            chart_zones.type = "bar"
            chart_zones.title = "Zones les plus affectées"
            chart_zones.x_axis.title = "Alertes"
            data_zones = Reference(ws_zones, min_col=2, min_row=1, max_row=zone_last_row)
            cats_zones = Reference(ws_zones, min_col=1, min_row=2, max_row=zone_last_row)
            chart_zones.add_data(data_zones, titles_from_data=True)
            chart_zones.set_categories(cats_zones)
            chart_zones.height = 12
            chart_zones.width = 15
            chart_zones.dLbls = DataLabelList()
            chart_zones.dLbls.showCatName = True
            chart_zones.dLbls.showVal = True
            ws_zones.add_chart(chart_zones, "E2")
        
        ws_zones.column_dimensions['A'].width = 20
        ws_zones.column_dimensions['B'].width = 12
        ws_zones.column_dimensions['C'].width = 12
        
        # ===== ONGLET 4: HISTORIQUE COMPLET =====
        ws_historique = wb.create_sheet("Historique", 3)
        
        ws_historique['A1'] = "Date"
        ws_historique['B1'] = "Heure"
        ws_historique['C1'] = "Zone"
        ws_historique['D1'] = "Risques"
        ws_historique['E1'] = "Détails"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_historique[f'{col}1'].font = Font(bold=True, color="FFFFFF")
            ws_historique[f'{col}1'].fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        
        hist_row = 2
        for alerte in sorted(alertes_semaine, key=lambda x: x.get("timestamp", ""), reverse=True):
            ws_historique[f'A{hist_row}'] = alerte.get("date", "")
            ws_historique[f'B{hist_row}'] = alerte.get("heure", "")
            ws_historique[f'C{hist_row}'] = _strip_report_emojis(alerte.get("zone", ""))
            ws_historique[f'D{hist_row}'] = _strip_report_emojis(alerte.get("risques", ""))
            ws_historique[f'E{hist_row}'] = f"T:{alerte.get('temp','?')} V:{alerte.get('wind','?')} P:{alerte.get('rain','?')}"
            
            if hist_row % 2 == 0:
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws_historique[f'{col}{hist_row}'].fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
            
            hist_row += 1
        
        ws_historique.column_dimensions['A'].width = 12
        ws_historique.column_dimensions['B'].width = 8
        ws_historique.column_dimensions['C'].width = 18
        ws_historique.column_dimensions['D'].width = 30
        ws_historique.column_dimensions['E'].width = 30
        
        os.makedirs(EXPORT_PATH, exist_ok=True)
        wb.save(RAPPORT_FILE)
        print(f"✅ Rapport Excel opérationnel généré : {RAPPORT_FILE}")
        return True
    
    except ImportError:
        print("⚠️ openpyxl manquant — installation...")
        import subprocess
        subprocess.run(["pip", "install", "openpyxl"], check=True)
        return generer_excel(alertes_semaine, stats)
    
    except Exception as e:
        print(f"❌ Erreur génération Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

def generer_excel_previsions(previsions):
    """Ajoute/écrit un fichier Excel séparé contenant les prévisions 5 jours."""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Prévisions 5J"
        ws.append(["Site", "Jour", "Tmin", "Tmax", "Pluie", "UV", "Risques"])
        for site, jours in previsions.items():
            for j in jours:
                ws.append([site, j['jour'], j['tmin'], j['tmax'], j['pluie'], j['uv'], j['risk']])
        path = os.path.join(EXPORT_PATH, "previsions_5j.xlsx")
        wb.save(path)
        print(f"✅ Prévisions 5j exportées : {path}")
        return path
    except Exception as e:
        print(f"❌ Erreur export prévisions: {e}")
        return None

def main(dry_run=False, force_send=False):
    print("=" * 60)
    print("📊 RAPPORT HEBDOMADAIRE DES RISQUES MÉTÉO")
    print("=" * 60)
    
    # Charger historique
    alertes, source_donnees = charger_historique()
    if REPORT_REQUIRE_RENDER and source_donnees in {"Fallback local", "API Render indisponible"}:
        raise RuntimeError(
            "Source Render obligatoire: impossible de générer le rapport sans API Render. "
            "Configurer JWT_SECRET/RENDER_URL ou autoriser explicitement ALLOW_LOCAL_REPORT_FALLBACK=true."
        )
    print(f"\n📁 Historique total: {len(alertes)} alertes")
    
    # Filtrer semaine précédente
    alertes_semaine = filtrer_semaine(alertes)
    lundi, dimanche = get_semaine_precedente()
    
    print(f"📅 Semaine analysée: {lundi} au {dimanche}")
    print(f"⚠️ Alertes de la semaine: {len(alertes_semaine)}")
    
    if len(alertes_semaine) == 0:
        print("✅ Aucune alerte cette semaine!")
        if not dry_run and not force_send:
            return
    
    # Générer statistiques
    stats = generer_statistiques(alertes_semaine)
    # Récupérer prévisions 5 jours pour affichage dans le rapport
    print("\n🔎 Récupération des prévisions 5 jours...")
    previsions_sites = fetch_previsions_5j(SITES, jours=5)
    previsions_voisins = fetch_previsions_5j(VOISINS, jours=5)
    
    # Générer Excel
    print("\n📄 Génération du rapport Excel...")
    generer_excel(alertes_semaine, stats)
    # ajouter un export séparé pour les prévisions (sites + voisins)
    all_previsions = {**previsions_sites, **previsions_voisins}
    previsions_file = generer_excel_previsions(all_previsions)
    
    # Envoyer par email
    print("\n📧 Envoi du rapport par email...")
    # joindre aussi le fichier prévisions si disponible
    envoyer_rapport_email(lundi, dimanche, stats, RAPPORT_FILE,
                         previsions_sites=previsions_sites,
                         previsions_voisins=previsions_voisins,
                         dry_run=dry_run,
                         source_donnees=source_donnees)
    
    print("\n" + "=" * 60)
    print("✅ Rapport hebdomadaire terminé")
    print("=" * 60)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Générer rapport hebdomadaire météo")
    parser.add_argument("--dry-run", action="store_true", help="Forcer génération et sauvegarder l'email localement sans l'envoyer")
    parser.add_argument("--force-send", action="store_true", help="Forcer envoi de l'email même sans alertes")
    args = parser.parse_args()
    main(dry_run=args.dry_run, force_send=args.force_send)
